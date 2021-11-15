"""
Run-me.py is the main file of the simulation. Run this file to run the simulation.
"""
from os import environ
import os

environ['PYGAME_HIDE_SUPPORT_PROMPT'] = '1'
import numpy as np
import pandas as pd
import networkx as nx
import matplotlib.pyplot as plt
import time as timer
import pygame as pg
from single_agent_planner import calc_heuristics
from visualization import map_initialization, map_running
from Aircraft import Aircraft
from distributed_planning import run_distributed_planner
# from cbs import run_CBS
import numpy.random as rnd
import time
from openpyxl import load_workbook

timertime = timer.time()
# %% SET SIMULATION PARAMETERS
# Input file names (used in import_layout) -> Do not change those unless you want to specify a new layout.
nodes_file = "nodes.xlsx"  # xlsx file with for each node: id, x_pos, y_pos, type
edges_file = "edges.xlsx"  # xlsx file with for each edge: from  (node), to (node), length



# Visualization (can also be changed)
plot_graph = False  # show graph representation in NetworkX
visualization_speed = 0.1  # set at 0.1 as default


# %%Function definitions
def import_layout(nodes_file, edges_file):
    """
    Imports layout information from xlsx files and converts this into dictionaries.
    INPUT:
        - nodes_file = xlsx file with node input data
        - edges_file = xlsx file with edge input data
    RETURNS:
        - nodes_dict = dictionary with nodes and node properties
        - edges_dict = dictionary with edges and edge properties
        - start_and_goal_locations = dictionary with node ids for arrival runways, departure runways and gates
    """
    gates_xy = []  # lst with (x,y) positions of gates
    rwy_dep_xy = []  # lst with (x,y) positions of entry points of departure runways
    rwy_arr_xy = []  # lst with (x,y) positions of exit points of arrival runways

    df_nodes = pd.read_excel(os.getcwd() + "/" + nodes_file)
    df_edges = pd.read_excel(os.getcwd() + "/" + edges_file)

    # Create nodes_dict from df_nodes
    nodes_dict = {}
    for i, row in df_nodes.iterrows():
        node_properties = {"id": row["id"],
                           "x_pos": row["x_pos"],
                           "y_pos": row["y_pos"],
                           "xy_pos": (row["x_pos"], row["y_pos"]),
                           "type": row["type"],
                           "neighbors": set()
                           }
        node_id = row["id"]
        nodes_dict[node_id] = node_properties

        # Add node type
        if row["type"] == "rwy_d":
            rwy_dep_xy.append((row["x_pos"], row["y_pos"]))
        elif row["type"] == "rwy_a":
            rwy_arr_xy.append((row["x_pos"], row["y_pos"]))
        elif row["type"] == "gate":
            gates_xy.append((row["x_pos"], row["y_pos"]))

    # Specify node ids of gates, departure runways and arrival runways in a dict
    start_and_goal_locations = {"gates": gates_xy,
                                "dep_rwy": rwy_dep_xy,
                                "arr_rwy": rwy_arr_xy}

    # Create edges_dict from df_edges
    edges_dict = {}
    for i, row in df_edges.iterrows():
        edge_id = (row["from"], row["to"])
        from_node = edge_id[0]
        to_node = edge_id[1]
        start_end_pos = (nodes_dict[from_node]["xy_pos"], nodes_dict[to_node]["xy_pos"])
        edge_properties = {"id": edge_id,
                           "from": row["from"],
                           "to": row["to"],
                           "length": row["length"],
                           "weight": row["length"],
                           "start_end_pos": start_end_pos
                           }
        edges_dict[edge_id] = edge_properties

    # Add neighbor nodes to nodes_dict based on edges between nodes
    for edge in edges_dict:
        from_node = edge[0]
        to_node = edge[1]
        nodes_dict[from_node]["neighbors"].add(to_node)

    return nodes_dict, edges_dict, start_and_goal_locations


def create_graph(nodes_dict, edges_dict, plot_graph=True):
    """
    Creates networkX graph based on nodes and edges and plots
    INPUT:
        - nodes_dict = dictionary with nodes and node properties
        - edges_dict = dictionary with edges annd edge properties
        - plot_graph = boolean (True/False) If True, function plots NetworkX graph. True by default.
    RETURNS:
        - graph = networkX graph object
    """

    graph = nx.DiGraph()  # create directed graph in NetworkX

    # Add nodes and edges to networkX graph
    for node in nodes_dict.keys():
        graph.add_node(node,
                       node_id=nodes_dict[node]["id"],
                       xy_pos=nodes_dict[node]["xy_pos"],
                       node_type=nodes_dict[node]["type"])

    for edge in edges_dict.keys():
        graph.add_edge(edge[0], edge[1],
                       edge_id=edge,
                       from_node=edges_dict[edge]["from"],
                       to_node=edges_dict[edge]["to"],
                       weight=edges_dict[edge]["length"])

    # Plot networkX graph
    if plot_graph:
        plt.figure()
        node_locations = nx.get_node_attributes(graph, 'xy_pos')
        nx.draw(graph, node_locations, with_labels=True, node_size=100, font_size=10)

    return graph


def scorecounter(aircraft_lst):  # Calculate score of simulation run
    # average waiting time
    wait_time = []
    for aircraft in aircraft_lst:
        wait_time.append(aircraft.waiting_time)

    avg_wait_time = np.average(wait_time)

    score = np.round(avg_wait_time, 4)

    return score


def inverse_nodes_dict():
    """
    Function to go from the (x,y) position back to the node. The return is a dictionary containing xy_positions and the
    ID of the respective nodes.

    This function is required to get the current node of the aircraft without for looping over a dictionary. Saves some
    computational time.

    Almost all code is copy pasted from create_graph. I adapted it to meet our needs.
    """
    nodes_file = "nodes.xlsx"
    df_nodes = pd.read_excel(os.getcwd() + "/" + nodes_file)

    inverse_nodes_dictionary = {}
    for i, row in df_nodes.iterrows():
        node_properties = {"id": row["id"],
                           "xy_pos": (row["x_pos"], row["y_pos"])}
        node_id = (row["x_pos"], row["y_pos"])
        inverse_nodes_dictionary[node_id] = node_properties
    return inverse_nodes_dictionary


# %% RUN SIMULATION
# =============================================================================
# 0. Initialization
# =============================================================================
nodes_dict, edges_dict, start_and_goal_locations = import_layout(nodes_file, edges_file)
inverse_nodes_dictionary = inverse_nodes_dict()
graph = create_graph(nodes_dict, edges_dict, plot_graph)
heuristics = calc_heuristics(graph, nodes_dict)
visualization = False  # pygame visualization


if visualization:
    map_properties = map_initialization(nodes_dict, edges_dict)  # visualization properties

# =============================================================================
# 1. While loop and visualization
# =============================================================================
# Parameters that can be changed:

simulation_time = 30
numb_of_aircraft = 40
running = True
escape_pressed = False
time_end = simulation_time + 5
dt = 0.1  # should be factor of 0.5 (0.5/dt should be integer)
t = 0
random = True  # True uses randomly generated aircraft, False generates 2 aircraft which collide at t = 5.0

planner = "Distributed"  # choose which planner to use (prioritized, CBS, Distributed)
priority = 'shortest_path'  # choose between 'first_come', 'shortest_path' or 'weighted'

def distributed_running(seed, running):
    rnd.seed(seed)
    N_max_cap = 0
    escape_pressed = False
    start_time = time.time()
    aircraft_lst = []  # List which can contain aircraft agents
    t = 0
    statement = True
    print("Simulation Started")
    while running:
        t = round(t, 2)

        # Check conditions for termination
        if t >= time_end or escape_pressed:
            running = False
            CPU_Time = timer.time() - start_time

            if CPU_Time > time_end:
                return False

            results['CPU_Time'] = round(CPU_Time, 3)
            cost = 0
            for ac in aircraft_lst:
                # results['Paths'].append(ac.path)
                temp_lst = []
                cost = cost + len(ac.locations)
                for i in ac.locations:
                    if i[0] not in temp_lst:
                        temp_lst.append(i[0])
                results['Total waiting time per aircraft'].append(len(ac.locations) - len(temp_lst))
            results['Total cost'] = cost
            results['Maximum waiting time'] = max(results['Total waiting time per aircraft'])
            results['Total waiting time'] = sum(results['Total waiting time per aircraft'])
            results['Average waiting time'] = round(results['Total waiting time']/len(aircraft_lst), 3)
            results['Number of generated aircraft'] = numb_of_aircraft
            results['Simulation time'] = simulation_time
            results['Number of ac'] = len(aircraft_lst)
            results['Max capacity'] = N_max_cap
            pg.quit()
            print("Simulation Stopped")
            return results

            # Visualization: Update map if visualization is true
        if visualization:
            current_states = {}  # Collect current states of all aircraft
            for ac in aircraft_lst:
                if ac.status == "taxiing":
                    current_states[ac.id] = {"ac_id": ac.id,
                                             "xy_pos": ac.position,
                                             "heading": ac.heading}
            escape_pressed = map_running(map_properties, current_states, t)
            timer.sleep(visualization_speed)

        # Spawning: aircraft are spawned on random gate/arrival nodes and are assigned a random departure/gate node
        gate_nodes = [97, 34, 35, 36, 98]
        arrival_nodes = [37, 38]
        departure_nodes = [1, 2]
        chosen_gate_nodes = [0, 0, 0, 0, 0]
        chosen_departure_nodes = [0, 0]


        if t == 1 and random == True:
            start_nodes_and_time = []

            # introduce random aircraft
            for i in range(numb_of_aircraft):
                counter = 0  # if multiple aircraft spawn at same place/time, counter goes up
                arrival_or_departure = rnd.choice(['A', 'D'])
                spawn_time = rnd.randint(1, simulation_time)

                if arrival_or_departure == 'A':
                    start_node = rnd.choice(arrival_nodes)
                    while [start_node, spawn_time] in start_nodes_and_time:
                        start_node = rnd.choice(arrival_nodes)
                        counter = counter + 1
                        if counter >= 2:  # first check other arrival node (of total 2)
                            spawn_time = rnd.randint(0, simulation_time)
                            counter = 0

                    goal_node = gate_nodes[chosen_gate_nodes.index(min(chosen_gate_nodes))]
                    chosen_gate_nodes[gate_nodes.index(goal_node)] += 1

                if arrival_or_departure == 'D':
                    start_node = rnd.choice(gate_nodes)
                    while [start_node, spawn_time] in start_nodes_and_time:
                        start_node = rnd.choice(gate_nodes)
                        counter = counter + 1
                        if counter >= 5:  # first check all other gate nodes (of total 5)
                            spawn_time = rnd.randint(0, simulation_time)
                            counter = 0

                    goal_node = departure_nodes[chosen_departure_nodes.index(min(chosen_departure_nodes))]
                    chosen_departure_nodes[departure_nodes.index(goal_node)] += 1

                ac = Aircraft(i, arrival_or_departure, start_node, goal_node, spawn_time, nodes_dict)
                aircraft_lst.append(ac)
                start_nodes_and_time.append([start_node, spawn_time])

        random = True
        # Spawn aircraft for this timestep (use for example a random process)
        if t == 1 and random == False:
            ac = Aircraft(17, 'A', 37, 36, t,
                          nodes_dict)  # As an example we will create one aicraft arriving at node 37 with the goal of reaching node 36
            ac1 = Aircraft(0, 'D', 36, 37, t,
                           nodes_dict)  # As an example we will create one aicraft arriving at node 36 with the goal of reaching node 37
            aircraft_lst.append(ac)
            aircraft_lst.append(ac1)

        elif planner == "Distributed":
            if t == 0:
                results = {'Max capacity': 0,  # Key: t
                           'Total waiting time': 0,
                           'Total waiting time per aircraft': [],
                           'Number of ac': 0,
                           'CPU_Time': 0}  # }
                constraints = {}
            if t % 0.5 == 0:
                statement = run_distributed_planner(aircraft_lst, nodes_dict, heuristics, t, constraints, inverse_nodes_dictionary)
        if statement == False:
            return False
        for ac in aircraft_lst:
            if ac.status == "taxiing":
                if t%0.5==0:
                    ac.locations.append((inverse_nodes_dictionary[ac.position]['id'], t))
                ac.move(dt, t)

        t = t + dt

        max_cap = 0
        for ac in aircraft_lst:
            if ac.status == 'taxiing':

                max_cap = max_cap + 1
        if max_cap > N_max_cap:
            N_max_cap = max_cap

total_result_dict = {}
counter = 0
seed = 0
cancelled_seeds = [] #[0, 1, 2, 3, 4, 5, 6, 15, 16, 17, 30, 32, 33, 39, 40, 42, 50, 53, 57, 58, 60, 61, 63, 65, 67, 74, 75, 77, 81, 82, 85, 87, 91, 92, 93, 100, 101, 108, 116, 118, 131]
while 100>counter:
    if visualization:
        map_properties = map_initialization(nodes_dict, edges_dict)  # visualization properties
    running = True
    if cancelled_seeds.count(seed) >= 1:
        seed = seed + 1
        continue
    print(seed)
    boolean = distributed_running(seed, running)
    if boolean == False:
        print('Cancelled seed ', seed)
        seed = seed + 1
        continue

    seed = seed + 1
    if type(boolean) == dict:
        boolean['seed'] = seed
        total_result_dict[counter] = boolean
    counter = counter + 1

    print(total_result_dict)
print('For 100 runs it took me: ', timer.time()-timertime)
file = "distributed.xlsx"
wb = load_workbook(file)
sheets = wb.sheetnames
sheet = wb[sheets[2]]

sheet.cell(row=1, column = 1).value = 'Simulation run'
sheet.cell(row=1, column = 2).value = 'Numb. of generated aircraft'
sheet.cell(row=1, column = 3).value = 'Total cost'
sheet.cell(row=1, column = 4).value = 'Total waiting time [s]'
sheet.cell(row=1, column = 5).value = 'Maximum delay'
sheet.cell(row=1, column = 6).value = 'Average waiting time [s]'
sheet.cell(row=1, column = 7).value = 'Maximum capacity'
sheet.cell(row=1, column = 8).value = 'CPU-time [s]'
sheet.cell(row=1, column = 9).value = 'Seed'

for i in range(len(total_result_dict)):
    sheet.cell(row=i+2, column = 1).value = i+1
    sheet.cell(row=i+2, column = 2).value = total_result_dict[i]['Number of ac']
    sheet.cell(row=i+2, column = 3).value = total_result_dict[i]['Total cost']
    sheet.cell(row=i+2, column = 4).value = total_result_dict[i]['Total waiting time']
    sheet.cell(row=i+2, column=5).value = total_result_dict[i]['Maximum waiting time']
    sheet.cell(row=i+2, column = 6).value = total_result_dict[i]['Average waiting time']
    sheet.cell(row=i+2, column = 7).value = total_result_dict[i]['Max capacity']
    sheet.cell(row=i+2, column = 8).value = total_result_dict[i]['CPU_Time']
    sheet.cell(row=i+2, column=9).value = total_result_dict[i]['seed']
wb.save(file)
